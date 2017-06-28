# -*- coding: utf-8 -*-
"""
Created on Sun Oct 16 20:45:07 2016

@author: Urban Rupnik
"""

import requests
import sys
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
from os import path, system
from openpyxl.chart import LineChart, Series, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from calendar import monthrange
from subprocess import call
import csv
import datetime


from nastavitve_zajemanje_tedni_cp import *

if(path.isdir(direktorij)==False):
	call(["mkdir", direktorij])
	call(["mkdir", direktorij+"Zajem"])
	call(["mkdir", direktorij+"Podatki"])

############################################
###############FUNKCIJE#####################
############################################


#funkcija, ki klicanemu datumu odšteje 10 minut in
#vrne nov datum v isti obliki
# vhodi: date -> string -> "dd.mm.YY (h)h:mm"
# izhodi: new_date -> string -> "dd.mm.YY (h)h:mm"
#			mesec -> int -> številka meseca novega datuma
# 			flag -> int -> zastavica, ki pove če se je zgodil preskok meseca


def rem_10(date):

	#poberemo dan, mesec, leto, uro in minuto iz stringa 'date'
	idx=date.find('.',0)
	dan=int(date[:idx])

	idx_old=idx
	idx=date.find('.',idx+1)
	mesec=int(date[idx_old+1:idx])
	
	idx_old=idx
	idx=date.find(' ',idx+1)
	leto=int(date[idx_old+1:idx])

	idx_old=idx
	idx=date.find(':',idx+1)
	ura=int(date[idx_old+1:idx])


	minuta=int(date[idx+1:])


	#prištejemo 10 minut in nastavimo zastavico za mesec na 0
	minuta-=10
	flag=0

	#naredimo popravke za datum(preskok v uri, datumu)
	if(minuta==-10):
		minuta=50
		ura-=1
		if(ura==-1):
			ura=23
			dan-=1
			if(dan==0):
				if(mesec==1):
					mesec=12
				else:
					mesec-=1
				day_num=monthrange(leto,mesec)[1]
				dan=day_num
				flag=1
				if(mesec==12):
					leto-=1
	#pripravimo string novega datuma				
	new_date=str(dan).zfill(2)+"."+str(mesec).zfill(2)+"."+str(leto)+" "+str(ura)+":"+str(minuta).zfill(2)
	
	return new_date, mesec, flag, leto;


############################################
###################TELO#####################
############################################


#število tednov v letu
st_tednov=52+1

#delanje backupa
try:
	call(["cp", direktorij+filename_tedni+".xlsx", direktorij+filename_tedni+"_cp.xlsx"])
except:
	pass
#priprava spremenljivk za strani v datoteki
#služi samo da ni kasnejših errorjev
ws=[None]*st_tednov
ws[0]=''

#Lista imen mesecev, da jih zapišemo na ime v vsaki strani
mesec1=['Januar','Februar','Marec','April', 
'Maj', 'Junij', 'Julij', 'Avgust', 
'September', 'Oktober', 'November', 'December']

#############################################
#################OBLIKOVANJE#################
#############################################

#določimo krepko pisavo za naslove, mrežo in centralno postavitev
ft=Font(bold=True)
bd1=Border(bottom=Side(border_style='thick', color='FF000000'))
al1=Alignment(horizontal='center', vertical='center')

#Ugotovimo, če datoteka že obstaja ali ne
#če obstaja, samo naložimo 'worksheete' v listo
if (path.isfile(direktorij+filename_tedni+".xlsx")):
	wb=load_workbook(direktorij+filename_tedni+".xlsx")
	for i in range(0,st_tednov):
		ws[i]=wb.get_sheet_by_name(str(i))


#drugače ustvarimo nov dokument, strani, ki jih poimenujemo
#in še drugo oblikovanje
else:
	wb=Workbook()
	for i in range(0,st_tednov):
		if (i==0):
			ws[0]=wb.active
			ws[0].title='1'
		else:
			ws[i]=wb.create_sheet(str(i))

		a1=ws[i]["A1"]
		b1=ws[i]["B1"]
		c1=ws[i]["C1"]
		d1=ws[i]["D1"]
		e1=ws[i]["E1"]
		f1=ws[i]["F1"]
		g1=ws[i]["G1"]

		a1.value="Dan"
		b1.value="Datum in ura"
		c1.value="Temperatura"
		d1.value="Vlažnost"
		e1.value="Padavine"
		f1.value="Padavine 12h"
		g1.value="Višina snežne odeje"

		col1=ws[i].column_dimensions['A']
		col2=ws[i].column_dimensions['B']
		col3=ws[i].column_dimensions['C']
		col4=ws[i].column_dimensions['D']
		col5=ws[i].column_dimensions['E']
		col6=ws[i].column_dimensions['F']
		col7=ws[i].column_dimensions['G']

		a1.font=ft
		b1.font=ft
		c1.font=ft
		d1.font=ft
		e1.font=ft
		f1.font=ft
		g1.font=ft

		a1.border=bd1
		b1.border=bd1
		c1.border=bd1
		d1.border=bd1
		e1.border=bd1
		f1.border=bd1
		g1.border=bd1

		col1.alignment=al1
		col2.alignment=al1
		col3.alignment=al1
		col4.alignment=al1
		col5.alignment=al1
		col6.alignment=al1
		col7.alignment=al1

		a1.alignment=al1
		b1.alignment=al1
		c1.alignment=al1
		d1.alignment=al1
		e1.alignment=al1
		f1.alignment=al1
		g1.alignment=al1

		col1.width = 10
		col2.width = 18
		col3.width = 14
		col4.width = 10
		col5.width = 10
		col6.width = 14
		col7.width = 23




#prvič shranimo datoteko
wb.save(direktorij+filename_tedni+".xlsx")

#sys.exit(1)


###################################
###########ZAJEM PODATKOV##########
###################################

#internetna stran iz katere zajemamo podatke o vremenu
#trenutno je nastavljena samo za to postajo, ker je če ne treba
#nastaviti druge indekse za stolpce
url = "http://meteo.arso.gov.si/uploads/probase/www/observ/surface/text/sl/observationAms_ZADLOG_history.html"
#file_out = "zadlog_vreme.csv"

#poskusimo pridobiti podatke iz strani
#če je poskus neuspešen, končamo aplikacijo
try:
    anas =requests.get(url)
except requests.exceptions.RequestException as e:  # This is the correct syntax
    print e
    sys.exit(1)

#print "h"

#uredimo kodiranje->lahko beremo tudi črke "č, ž, š"
anas.encoding='utf-8'
#pretvorba v tekst
text=anas.text
#ustvarimo parser?
soup = BeautifulSoup(text, 'html.parser')

#######################################
#########BRANJE ŽE OBSTOJEČEGA#########
#######################################

#pridobivanje zadnjega podatka v datoteki
#najprej po mesecih -> začnemo iz zadnje strani
i=st_tednov-1-1
while i>=0:
	#če je druga vrstica v stolpcu za datume prazna, 
	#je mesec prazen
	last_old_month = ws[i]['B2'].value
	if(last_old_month != None):
		break
	else:
		i-=1

if(i==-1):
	mesec=0
else:
	mesec=i
#nato po podatkih -> začnemo iz 2. vrstice		
j=2
while True:
	last_old_value=ws[i]['B'+str(j)].value
	#print last_old_value
	if (last_old_value != None):
		j+=1
	elif ((last_old_value == None) and (j==2)):
		j-=1 
		last_old_value="01."+str(mesec+1).zfill(2)+"."+leto_str+" 0:00"
		break
	else:
		j-=1 
		#preberemo datum in uro zadnjega podatka
		last_old_value=ws[i]['B'+str(j)].value
		break
#print last_old_value;

#ustvarimo listo podatkov
data=[]
#preparsanje tabele
table = soup.table

##############################
########BRANJE PODATKOV#######
##############################

for index1, tr in enumerate(table.find_all("tr")[1:]):  # skip first row
	tds = tr.find_all("td")
	time = tds[0].text #  take the one without "Nedelja, "
	idx = time.find(",")  # find index and get day, date and time separately
	day = time[:idx]
	date = time[idx+2:idx+12]
	idx1=time.find(" ",idx+10)
	idx=time.find(":")
	time1 = time[idx1+1:idx+3]
	temperature_precise = tds[10].text #branje temperature z eno decimalko
	temparature = tds[11].text #branje temperature, zaokroženo
	humidity_precise = tds[12].text #branje vlažnosti, z eno decimalko
	humidity 	= tds[13].text	#branje vlažnosti, brez decimalke
	rainfall = tds[24].text	#branje padavin z eno decimalko
	rainfall_12h = tds[26].text #branje padavin v 12h
	snow=tds[32].text
	new_old_value=date+ ' ' + time1

	#če je datum podatka enak zadnjemu vpisanemu,
	#potem ne beremo več
	if(index1 == 0):
		old_value=new_old_value
		c_tmp=int(date[6:10])
	if(new_old_value == last_old_value):
		out_flag=1
		break;

	temp_value,a,b,c=rem_10(old_value)

	if(c<c_tmp):
		file = open("nastavitve_zajemanje_tedni_cp.py", "w")
		file.write("leto_str='"+str(c_tmp)+"'\n") 
		file.write("direktorij='/home/urban/Documents/Projekti/Vreme/"+str(c_tmp)+"_cp/'\n")
		file.write("filename_tedni='Zajem/vreme_tedni_'+leto_str\n")
		file.write("filename='Zajem/vreme_'+leto_str\n")
		file.close()	
	else:
		pass
	


	if(int(leto_str)==c):
		#print temp_value + ' ' + new_old_value
		#print str(len(temp_value)) + ' ' + str(len(new_old_value))
		if (temp_value==new_old_value):
			pass
		#	print 'OK'
		ctr=0
		while (1):
			if(old_value==new_old_value):
			#	print 's1'
				break
			elif(temp_value==new_old_value):
				#print 's2'
				break
			else:
				ctr +=1
				#print temp_value
				data.append((temp_value, '','','', '',''))
				temp_value,a,b=rem_10(temp_value)

			if (ctr==100000):
				print "exit"
				sys.exit(1)
				break

		old_value=new_old_value
		out_flag=0
		#dodamo vrstico podatkov k obstoječim
		data.append((date+ ' ' + time1, day, temperature_precise, humidity_precise, rainfall, rainfall_12h,snow))

#print out_flag
#print last_old_value
#print len(last_old_value)

if(out_flag==0):
	ctr = 0
	temp_value,a,b,c=rem_10(old_value)
	while (1):
		if(temp_value==last_old_value):
			#print "h14"
			break
		else:
			ctr +=1
			data.append((temp_value, '','','', '',''))
			temp_value,a,b,c=rem_10(temp_value)

		if (ctr==100000):
			print "exit"
			sys.exit(1)
			break
#obrnemo listo, da imamo najstarejše podatke na vrhu
data.reverse();


#ker lahko pride zaradi prenehanja delovanja postaje ali serverj do tega,
# da novih podatkov ni, ali pa vmes ne oddajajo, zarati risanja grafov
# vrinemo 'slepe' podatke, ki imajo samo datum, vrednosti pa ne

#!!!!!!!! še ni testirano in ni pravilno!!!
#while(1):
	#iz funkcije preberemo nov datum, mesec, in če je bil preskok meseca
#	last_old_value1, month1, zastavca=add_10(last_old_value)
	
	#če sploh ni novih podatkov, takrat ne napišemo nič
#	if(last_old_value==new_old_value):
#		break
	#če so novi podatki, ampak tudi ni luknje v datumu, ne pišemo nič
#	elif(last_old_value1==new_old_value):
#		break
#	else:
#		if(zastavca==1):
#			j=1

#		ws[month1-1]['B'+str(j+1)]=last_old_value1
#		last_old_value=last_old_value1



empty_flag=0
temp_index=0
week = None

#vsako vrednost v vsakem podatku zapišemo v datoteko
for index, sample in enumerate(data):
	for idx, x in enumerate(sample):
		#Pogledamo, če slučajno nimamo vrednosti za podatek
		#zaradi potrebe risanja, ker če zapišemo prazen string
		#imamo probleme
		if (x==''):
			empty_flag=1
		else:
			empty_flag=0

		if(idx == 0):
			dotidx1=x.find('.')
			dotidx2=x.find('.', dotidx1+1)
			dotidx3=x.find(' ', dotidx2+1)

			old_week=week
			week=datetime.date(int(x[dotidx2+1:dotidx3]),int(x[dotidx1+1:dotidx2]), int(x[0:dotidx1])).isocalendar()[1]-1
			
						
			if(week==52 and int(x[dotidx1+1:dotidx2])==1):
				week=0
			
			#month = int(x[dotidx1+1:dotidx2])-1
			if (old_week != week and old_week != None):
				temp_index = index
				j=1
			ws[week]['B'+str(index-temp_index+j+1)]=x
		elif(idx == 1):
			ws[week]['A'+str(index-temp_index+j+1)]=x
		elif(idx==2):
			if (empty_flag==1):
				ws[week]['C'+str(index-temp_index+j+1)]=None
			else:	
				ws[week]['C'+str(index-temp_index+j+1)]=float(x)
		elif(idx==3):
			if (empty_flag==1):
				ws[week]['D'+str(index-temp_index+j+1)]=None
			else:
				ws[week]['D'+str(index-temp_index+j+1)]= float(x)
		elif(idx==4):
			if (empty_flag==1):
				ws[week]['E'+str(index-temp_index+j+1)]=None
			else:
				ws[week]['E'+str(index-temp_index+j+1)]= float(x)
		elif(idx==5):
			if (empty_flag==1):
				ws[week]['F'+str(index-temp_index+j+1)]=None
			else:
				ws[week]['F'+str(index-temp_index+j+1)]= float(x)
		elif(idx==6):
			if (empty_flag==1):
				ws[week]['G'+str(index-temp_index+j+1)]=None
			else:
				ws[week]['G'+str(index-temp_index+j+1)]= float(x)			
#shranimo datoteko
wb.save(direktorij+filename_tedni+".xlsx")




for i in range(0,st_tednov):
	with open(direktorij+'Podatki/'+str(i)+'.csv', 'wb') as f:
		c = csv.writer(f)
		for idxrow, r in enumerate(ws[i].rows):
			if idxrow != 0:
				c.writerow([cell.value for cell in r[1:]])	
		