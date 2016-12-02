from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import time
import datetime
import matplotlib.pyplot as plt 
import datetime
import sys
from numpy import arange
from matplotlib.dates import DayLocator, HourLocator, DateFormatter

########################################
#############BRANJE PODATKOV############
########################################

#datoteka iz katere beremo
filename = '/home/rupnik/Documents/Projekti/Vreme/vreme_2016.xlsx'

#preberemo mesec in leto trenutnega datuma
#mesec zmanjsamo za 1, ker je mesec nazaj napolnjen s podatki
m=int(time.strftime("%m"))-1
#popravek za risanje Decembra (1->12)
if(m==0):
	m=12
y=int(time.strftime("%Y"))

#Lista imen mesecev, da jih zapišemo na ime v vsaki strani
mesec=['Januar','Februar','Marec','April', 
'Maj', 'Junij', 'Julij', 'Avgust', 
'September', 'Oktober', 'November', 'December']


#v spomin nalozimo datoteko in ustrezno stran
wb=load_workbook(filename)
ws=wb.worksheets[m-1]

#deklaracija list
x_data=[]
temp=[]
hum=[]
rain=[]
j=[]

date2=0
date=[]

#preberemo podatke in ustvarimo listo datumov, drugace ne izrise!
i=2
while(ws['B'+str(i)].value!=None):
	x_data.append(ws['B'+str(i)].value)
	temp.append(ws['C'+str(i)].value)
	hum.append(ws['D'+str(i)].value)
	rain.append(ws['E'+str(i)].value)

	date1=[]
	#day,month,year, hour,minute
	idx=x_data[i-2].find('.',0)
	date2=x_data[i-2][:idx]
	date1.append(int(date2))

	idx_old=idx
	idx=x_data[i-2].find('.',idx+1)
	date2=x_data[i-2][idx_old+1:idx]
	date1.append(int(date2))
	
	idx_old=idx
	idx=x_data[i-2].find(' ',idx+1)
	date2=x_data[i-2][idx_old+1:idx]
	date1.append(int(date2))

	idx_old=idx
	idx=x_data[i-2].find(':',idx+1)
	date2=x_data[i-2][idx_old+1:idx]
	date1.append(int(date2))


	date2=x_data[i-2][idx+1:]
	date1.append(int(date2))


	

	date.append(datetime.datetime(date1[2], date1[1], date1[0], date1[3], date1[4]))


	i+=1

#kontrola za prazno stran
if i==2:
	print ("Prazna stran za mesec %s v letu %d" % (mesec[m-1], y))
	sys.exit(1)

#risanje grafa temperature v odvisnosti od datuma
fig1, ax = plt.subplots()
ax.plot_date(date, temp, 'r-')#risanje

ax.xaxis.set_major_locator(DayLocator())#risanje glavnih oznak (x os)
ax.xaxis.set_minor_locator(HourLocator(arange(0, 25, 6))) #risanje pomoznih oznak (x os)
ax.xaxis.set_major_formatter(DateFormatter('%d.%m.%Y'))#format oznacitve

ax.set_title('Temperatura - %s %d' % (mesec[m-1],y))
ax.set_xlabel('Datum')
ax.set_ylabel('Temperatura [$^{\circ}$C]')

plt.ylim((-35,35))
plt.grid(True)

ax.fmt_xdata = DateFormatter('%d.%m.%Y %H:%M:%S')
fig1.autofmt_xdate() #popravek za lepsi zapis (postrani)

#shranitev strani
plt.savefig('/home/rupnik/Documents/Projekti/Vreme/temp_'+str(m)+'_'+str(y)+'_test.png')


#ponovimo se za vlaznost
fig2, ax = plt.subplots()

ax.plot_date(date, hum, 'g-')

ax.xaxis.set_major_locator(DayLocator())
ax.xaxis.set_minor_locator(HourLocator(arange(0, 25, 6)))

ax.xaxis.set_major_formatter(DateFormatter('%d.%m.%Y'))#format oznacitve

ax.set_title('Vlaznost - %s %d' % (mesec[m-1],y))
ax.set_xlabel('Datum')
ax.set_ylabel('Vlaznost [%]')

plt.ylim((40,105))
plt.grid(True)

ax.fmt_xdata = DateFormatter('%d.%m.%Y %H:%M:%S')
fig2.autofmt_xdate()
plt.savefig('/home/rupnik/Documents/Projekti/Vreme/hum_'+str(m)+'_'+str(y)+'_test.png')

#in padavine
fig3, ax = plt.subplots()

ax.plot_date(date, rain, 'b-')

ax.xaxis.set_major_locator(DayLocator())
ax.xaxis.set_minor_locator(HourLocator(arange(0, 25, 6)))
ax.xaxis.set_major_formatter(DateFormatter('%d.%m.%Y'))#format oznacitve

ax.set_title('Padavine - %s %d' % (mesec[m-1],y))
ax.set_xlabel('Datum')
ax.set_ylabel('Padavine [mm]')

plt.ylim((-0.5,7))
plt.grid(True)

ax.fmt_xdata = DateFormatter('%d.%m.%Y %H:%M:%S')
fig3.autofmt_xdate()
plt.savefig('/home/rupnik/Documents/Projekti/Vreme/rain_'+str(m)+'_'+str(y)+'_test.png')

#prikaz
plt.show()
