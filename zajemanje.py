# -*- coding: utf-8 -*-
"""
Created on Sun Oct 16 20:45:07 2016

@author: Urban Rupnik
"""

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
from os import path


filename = 'test1.xlsx'
ws=[None]*12
ws[0]=''

mesec=['Januar','Februar','Marec','April', 
'Maj', 'Junij', 'Julij', 'Avgust', 
'September', 'Oktober', 'November', 'December']

ft=Font(bold=True)
bd1=Border(bottom=Side(border_style='thick', color='FF000000'))
al1=Alignment(horizontal='center', vertical='center')

if (path.isfile(filename)):
	wb=load_workbook(filename)
	for i in range(0,12):
		ws[i]=wb.get_sheet_by_name(mesec[i])

else:
	wb=Workbook()
	for i in range(0,12):
		if (i==0):
			ws[0]=wb.active
			ws[0].title='Januar'
		else:
			ws[i]=wb.create_sheet(mesec[i])

		a1=ws[i]["A1"]
		b1=ws[i]["B1"]
		c1=ws[i]["C1"]
		d1=ws[i]["D1"]
		e1=ws[i]["E1"]
		f1=ws[i]["F1"]

		a1.value="Dan"
		b1.value="Datum in ura"
		c1.value="Temperatura"
		d1.value="Vlažnost"
		e1.value="Padavine"
		f1.value="Padavine 12h"

		col1=ws[i].column_dimensions['A']
		col2=ws[i].column_dimensions['B']
		col3=ws[i].column_dimensions['C']
		col4=ws[i].column_dimensions['D']
		col5=ws[i].column_dimensions['E']
		col6=ws[i].column_dimensions['F']

		a1.font=ft
		b1.font=ft
		c1.font=ft
		d1.font=ft
		e1.font=ft
		f1.font=ft

		a1.border=bd1
		b1.border=bd1
		c1.border=bd1
		d1.border=bd1
		e1.border=bd1
		f1.border=bd1

		col1.alignment=al1
		col2.alignment=al1
		col3.alignment=al1
		col4.alignment=al1
		col5.alignment=al1
		col6.alignment=al1

		a1.alignment=al1
		b1.alignment=al1
		c1.alignment=al1
		d1.alignment=al1
		e1.alignment=al1
		f1.alignment=al1

		col1.width = 10
		col2.width = 18
		col3.width = 14
		col4.width = 10
		col5.width = 10
		col6.width = 14





wb.save(filename)

url = "http://meteo.arso.gov.si/uploads/probase/www/observ/surface/text/sl/observationAms_ZADLOG_history.html"
file_out = "zadlog_vreme.csv"

anas =requests.get(url)
anas.encoding='utf-8'
text=anas.text
soup = BeautifulSoup(text, 'html.parser')


i=11
while i>=0:
	last_old_month = ws[i]['A2'].value
	if(last_old_month != None):
		break
	else:
		i-=1
		
j=2
while True:
	last_old_value=ws[i]['B'+str(j)].value
	if (last_old_value != None):
		j+=1
	else:
		j-=1 
		last_old_value=ws[i]['B'+str(j)].value
		break


data=[]
table = soup.table


for index1, tr in enumerate(table.find_all("tr")[1:]):  # skip first row
    tds = tr.find_all("td")
    time = tds[0].text #  take the one without "Nedelja, "
    idx = time.find(",")  # find index and get day, date and time separately
    day = time[:idx]
    #print(day)
    date = time[idx+2:idx+12]
    #print(date)
    time1 = time[idx+13:idx+18]
    #print(time1)
    temperature_precise = tds[10].text
    temparature = tds[11].text
    humidity_precise = tds[12].text
    humidity = tds[13].text
    rainfall = tds[24].text
    rainfall_12h = tds[26].text
    if(date+ ' ' + time1 == last_old_value):
		break

    data.append((date+ ' ' + time1, day, temperature_precise, humidity_precise, rainfall, rainfall_12h))
data.reverse();

empty_flag=0
temp_index=0
month = None

for index, sample in enumerate(data):
	for idx, x in enumerate(sample):
		if (x==''):
			empty_flag=1
		else:
			empty_flag=0

		if(idx == 0):
			dotidx1=x.find('.')
			dotidx2=x.find('.', dotidx1+1)
			old_month=month
			month = int(x[dotidx1+1:dotidx2])-1
			if (old_month != month and old_month != None):
				temp_index = index
				j=1
			ws[month]['B'+str(index-temp_index+j+1)]=x
		elif(idx == 1):
			ws[month]['A'+str(index-temp_index+j+1)]=x
		elif(idx==2):
			if (empty_flag==1):
				ws[month]['C'+str(index-temp_index+j+1)]=None
			else:	
				ws[month]['C'+str(index-temp_index+j+1)]=float(x)
		elif(idx==3):
			if (empty_flag==1):
				ws[month]['D'+str(index-temp_index+j+1)]=None
			else:
				ws[month]['D'+str(index-temp_index+j+1)]= float(x)
		elif(idx==4):
			if (empty_flag==1):
				ws[month]['E'+str(index-temp_index+j+1)]=None
			else:
				ws[month]['E'+str(index-temp_index+j+1)]= float(x)
		elif(idx==5):
			if (empty_flag==1):
				ws[month]['F'+str(index-temp_index+j+1)]=None
			else:
				ws[month]['F'+str(index-temp_index+j+1)]= float(x)
#		elif(idx==6):
#			ws[month]['G'+str(index+j+1)]=float(x)
#		elif(idx==7):
#			ws[month]['H'+str(index+2)]=x
#print(data)
#for x in data:
#	print(x)


wb.save(filename)
